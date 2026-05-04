from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import Group, User
from django.db import transaction
from django.db.models import Avg, Count, Max, Q
from django.http import HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse, reverse_lazy
from django.utils.text import slugify
from django.views.generic import CreateView, DeleteView, DetailView, ListView, TemplateView, UpdateView
from openpyxl import load_workbook

from core.permissions import AdminRequiredMixin
from grading.models import Attendance, Course, Grade, Lesson
from users.forms import ChatDialogForm, ChatMessageForm, ParentForm, StudentForm, TeacherForm, StudentImportForm
from users.models import ChatDialog, ChatMessage, Parent, Student, Teacher


def _contains_casefold(haystack, needle):
    return needle.casefold() in (haystack or '').casefold()


class TeacherListView(AdminRequiredMixin, ListView):
    model = Teacher
    template_name = 'users/teacher_list.html'
    paginate_by = 20

    def get_queryset(self):
        qs = Teacher.objects.select_related('user', 'department')
        q = (self.request.GET.get('q') or '').strip()
        if q:
            matched_ids = []
            for obj in qs:
                fio = (obj.user.get_full_name() or '').strip()
                if (
                    _contains_casefold(fio, q)
                    or _contains_casefold(obj.user.first_name, q)
                    or _contains_casefold(obj.user.last_name, q)
                    or _contains_casefold(obj.user.username, q)
                ):
                    matched_ids.append(obj.id)
            qs = qs.filter(id__in=matched_ids)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = (self.request.GET.get('q') or '').strip()
        return context


class TeacherCreateView(AdminRequiredMixin, CreateView):
    model = Teacher
    form_class = TeacherForm
    template_name = 'core/form.html'
    success_url = reverse_lazy('teacher_list')


class TeacherUpdateView(AdminRequiredMixin, UpdateView):
    model = Teacher
    form_class = TeacherForm
    template_name = 'core/form.html'
    success_url = reverse_lazy('teacher_list')


class TeacherDeleteView(AdminRequiredMixin, DeleteView):
    model = Teacher
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('teacher_list')


class StudentListView(AdminRequiredMixin, ListView):
    model = Student
    template_name = 'users/student_list.html'
    paginate_by = 20

    def get_queryset(self):
        qs = Student.objects.select_related('user', 'group')
        q = (self.request.GET.get('q') or '').strip()
        if q:
            matched_ids = []
            for obj in qs:
                fio = (obj.user.get_full_name() or '').strip()
                if (
                    _contains_casefold(fio, q)
                    or _contains_casefold(obj.user.first_name, q)
                    or _contains_casefold(obj.user.last_name, q)
                    or _contains_casefold(obj.user.username, q)
                ):
                    matched_ids.append(obj.id)
            qs = qs.filter(id__in=matched_ids)
        group = self.request.GET.get('group')
        if group:
            qs = qs.filter(group_id=group)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = (self.request.GET.get('q') or '').strip()
        return context


class StudentCreateView(AdminRequiredMixin, CreateView):
    model = Student
    form_class = StudentForm
    template_name = 'core/form.html'
    success_url = reverse_lazy('student_list')


class StudentUpdateView(AdminRequiredMixin, UpdateView):
    model = Student
    form_class = StudentForm
    template_name = 'core/form.html'
    success_url = reverse_lazy('student_list')


class StudentDeleteView(AdminRequiredMixin, DeleteView):
    model = Student
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('student_list')


class ParentListView(AdminRequiredMixin, ListView):
    model = Parent
    template_name = 'users/parent_list.html'
    paginate_by = 20

    def get_queryset(self):
        qs = Parent.objects.select_related('user').prefetch_related('children__user', 'children__group')
        q = (self.request.GET.get('q') or '').strip()
        if q:
            matched_ids = []
            for obj in qs:
                fio = (obj.user.get_full_name() or '').strip()
                children = ' '.join(str(child) for child in obj.children.all())
                if (
                    _contains_casefold(fio, q)
                    or _contains_casefold(obj.user.first_name, q)
                    or _contains_casefold(obj.user.last_name, q)
                    or _contains_casefold(obj.user.username, q)
                    or _contains_casefold(children, q)
                ):
                    matched_ids.append(obj.id)
            qs = qs.filter(id__in=matched_ids)
        return qs

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['q'] = (self.request.GET.get('q') or '').strip()
        return context


class ParentCreateView(AdminRequiredMixin, CreateView):
    model = Parent
    form_class = ParentForm
    template_name = 'core/form.html'
    success_url = reverse_lazy('parent_list')


class ParentUpdateView(AdminRequiredMixin, UpdateView):
    model = Parent
    form_class = ParentForm
    template_name = 'core/form.html'
    success_url = reverse_lazy('parent_list')


class ParentDeleteView(AdminRequiredMixin, DeleteView):
    model = Parent
    template_name = 'core/confirm_delete.html'
    success_url = reverse_lazy('parent_list')


class ParentDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'users/parent_dashboard.html'

    def dispatch(self, request, *args, **kwargs):
        if not _is_parent(request.user):
            return HttpResponseForbidden('Нет доступа к кабинету родителя')
        self.parent = get_object_or_404(Parent.objects.prefetch_related('children__user', 'children__group'), user=request.user)
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        children = self.parent.children.select_related('user', 'group').order_by('user__last_name', 'user__first_name')
        selected_id = self.request.GET.get('student')
        selected_student = children.filter(id=selected_id).first() if selected_id else children.first()

        context['parent'] = self.parent
        context['children'] = children
        context['selected_student'] = selected_student

        if not selected_student:
            context.update(
                {
                    'courses': [],
                    'attendance_percent': 0,
                    'attendance_total': 0,
                    'grades_count': 0,
                    'student_avg': None,
                    'subject_rows': [],
                    'attendance_rows': [],
                }
            )
            return context

        courses = (
            Course.objects.filter(group=selected_student.group)
            .select_related('subject', 'teacher__user')
            .annotate(
                avg_grade=Avg('grades__value', filter=Q(grades__student=selected_student)),
                grades_count=Count('grades', filter=Q(grades__student=selected_student)),
                attendance_total=Count('lessons__attendances', filter=Q(lessons__attendances__student=selected_student)),
                attendance_present=Count(
                    'lessons__attendances',
                    filter=Q(
                        lessons__attendances__student=selected_student,
                        lessons__attendances__status__in=[Attendance.Status.PRESENT, Attendance.Status.LATE],
                    ),
                ),
            )
            .order_by('subject__name')
        )
        grades = Grade.objects.filter(student=selected_student)
        attendance_qs = Attendance.objects.filter(student=selected_student)
        attendance_total = attendance_qs.count()
        attendance_present = attendance_qs.filter(status__in=[Attendance.Status.PRESENT, Attendance.Status.LATE]).count()

        subject_rows = []
        attendance_rows = []
        for course in courses:
            subject_rows.append(
                {
                    'name': course.subject.name,
                    'avg_grade': float(course.avg_grade) if course.avg_grade is not None else None,
                    'grades_count': course.grades_count,
                }
            )
            attendance_rows.append(
                {
                    'name': course.subject.name,
                    'percent': round((course.attendance_present / course.attendance_total) * 100, 2)
                    if course.attendance_total
                    else 0,
                    'total': course.attendance_total,
                }
            )

        context.update(
            {
                'courses': courses,
                'attendance_percent': round((attendance_present / attendance_total) * 100, 2) if attendance_total else 0,
                'attendance_total': attendance_total,
                'grades_count': grades.count(),
                'student_avg': grades.aggregate(avg=Avg('value'))['avg'],
                'subject_rows': subject_rows,
                'attendance_rows': attendance_rows,
            }
        )
        return context


class ChatListView(LoginRequiredMixin, ListView):
    model = ChatDialog
    template_name = 'users/chat_list.html'
    paginate_by = 20

    def get_queryset(self):
        user = self.request.user
        qs = ChatDialog.objects.select_related(
            'student__user',
            'teacher__user',
            'related_grade__course__subject',
        ).prefetch_related('messages')
        if _is_admin(user):
            pass
        elif Teacher.objects.filter(user=user).exists():
            qs = qs.filter(teacher__user=user)
        elif Student.objects.filter(user=user).exists():
            qs = qs.filter(student__user=user)
        else:
            qs = qs.none()

        q = (self.request.GET.get('q') or '').strip()
        if q:
            qs = qs.filter(
                Q(title__icontains=q)
                | Q(messages__message__icontains=q)
                | Q(student__user__first_name__icontains=q)
                | Q(student__user__last_name__icontains=q)
                | Q(teacher__user__first_name__icontains=q)
                | Q(teacher__user__last_name__icontains=q)
            ).distinct()
        return qs.annotate(last_message_at=Max('messages__created_at')).order_by('-updated_at', '-last_message_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        rows = []
        for chat in context['object_list']:
            last_message = chat.messages.order_by('-created_at').first()
            unread_count = chat.messages.filter(is_read=False).exclude(sender=self.request.user).count()
            companion = chat.teacher if chat.student.user_id == self.request.user.id else chat.student
            rows.append(
                {
                    'chat': chat,
                    'last_message': last_message,
                    'unread_count': unread_count,
                    'companion': companion,
                }
            )
        context['chat_rows'] = rows
        context['q'] = (self.request.GET.get('q') or '').strip()
        context['can_start_chat'] = Student.objects.filter(user=self.request.user).exists()
        return context


class ChatCreateView(LoginRequiredMixin, TemplateView):
    template_name = 'users/chat_form.html'

    def dispatch(self, request, *args, **kwargs):
        self.student = Student.objects.filter(user=request.user).select_related('group').first()
        if not self.student:
            return HttpResponseForbidden('Создавать диалоги может только студент')
        grade_id = request.GET.get('grade') or request.POST.get('grade')
        self.grade = None
        if grade_id:
            self.grade = get_object_or_404(Grade.objects.select_related('student', 'course__teacher__user', 'course__subject'), pk=grade_id)
            if self.grade.student_id != self.student.id:
                return HttpResponseForbidden('Нет доступа к этой оценке')
        return super().dispatch(request, *args, **kwargs)

    def initial_teacher(self):
        teacher_id = self.request.GET.get('teacher') or self.request.POST.get('teacher')
        if self.grade:
            return self.grade.course.teacher
        if teacher_id:
            return Teacher.objects.filter(pk=teacher_id).first()
        return None

    def get_form(self):
        data = self.request.POST or None
        form = ChatDialogForm(data, student=self.student, grade=self.grade)
        teacher = self.initial_teacher()
        if teacher and not data:
            form.fields['teacher'].initial = teacher.pk
        return form

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['form'] = kwargs.get('form') or self.get_form()
        context['grade_context'] = _grade_context(self.grade)
        context['grade'] = self.grade
        return context

    def post(self, request, *args, **kwargs):
        raw_teacher_id = request.POST.get('teacher')
        if raw_teacher_id and not self.grade:
            requested_teacher = Teacher.objects.filter(pk=raw_teacher_id).first()
            if requested_teacher and not _student_can_write_teacher(self.student, requested_teacher):
                return HttpResponseForbidden('Можно писать только преподавателям своих курсов')
        form = self.get_form()
        if not form.is_valid():
            return render(request, self.template_name, self.get_context_data(form=form))
        teacher = form.cleaned_data['teacher']
        if not _student_can_write_teacher(self.student, teacher):
            return HttpResponseForbidden('Можно писать только преподавателям своих курсов')
        with transaction.atomic():
            chat = ChatDialog.objects.create(
                student=self.student,
                teacher=teacher,
                related_grade=self.grade,
                title=form.cleaned_data['title'],
            )
            ChatMessage.objects.create(
                chat=chat,
                sender=request.user,
                sender_role=ChatMessage.SenderRole.STUDENT,
                message=form.cleaned_data['message'],
                is_read=False,
            )
        messages.success(request, 'Диалог создан.')
        return redirect('chat_detail', pk=chat.pk)


class ChatDetailView(LoginRequiredMixin, DetailView):
    model = ChatDialog
    template_name = 'users/chat_detail.html'

    def get_queryset(self):
        return ChatDialog.objects.select_related(
            'student__user',
            'teacher__user',
            'related_grade__student__user',
            'related_grade__course__subject',
            'related_grade__course__teacher__user',
        ).prefetch_related('messages__sender')

    def dispatch(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not _can_access_chat(request.user, self.object):
            return HttpResponseForbidden('Нет доступа к диалогу')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        chat = self.object
        chat.messages.filter(is_read=False).exclude(sender=self.request.user).update(is_read=True)
        context = super().get_context_data(**kwargs)
        context['form'] = kwargs.get('form') or ChatMessageForm()
        context['messages_list'] = chat.messages.select_related('sender')
        context['grade_context'] = _grade_context(chat.related_grade)
        context['companion'] = chat.teacher if chat.student.user_id == self.request.user.id else chat.student
        return context

    def post(self, request, *args, **kwargs):
        chat = self.object
        form = ChatMessageForm(request.POST, request.FILES)
        if not form.is_valid():
            return render(request, self.template_name, self.get_context_data(form=form))
        sender_role = _chat_role(request.user)
        if not sender_role:
            return HttpResponseForbidden('Нет роли для отправки сообщений')
        message = form.save(commit=False)
        message.chat = chat
        message.sender = request.user
        message.sender_role = sender_role
        message.is_read = False
        message.save()
        chat.save(update_fields=['updated_at'])
        return redirect('chat_detail', pk=chat.pk)


class StudentDetailView(LoginRequiredMixin, DetailView):
    model = Student
    template_name = 'users/student_detail.html'

    def get_queryset(self):
        return Student.objects.select_related('user', 'group')

    def dispatch(self, request, *args, **kwargs):
        student = self.get_object()
        if not _can_access_student_profile(request.user, student):
            return HttpResponseForbidden('Нет доступа к профилю студента')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.object
        context['subjects'] = (
            Course.objects.filter(group=student.group)
            .select_related('subject', 'teacher__user')
            .annotate(
                avg_grade=Avg('grades__value', filter=Q(grades__student=student)),
                grades_count=Count('grades', filter=Q(grades__student=student)),
            )
            .order_by('subject__name')
        )
        return context


class StudentCourseJournalView(LoginRequiredMixin, TemplateView):
    template_name = 'users/student_course_journal.html'

    def dispatch(self, request, *args, **kwargs):
        self.student = get_object_or_404(Student.objects.select_related('user', 'group'), pk=self.kwargs['student_id'])
        self.course = get_object_or_404(
            Course.objects.select_related('subject', 'group', 'teacher__user'),
            pk=self.kwargs['course_id'],
            group=self.student.group,
        )
        if not _can_access_student_profile(request.user, self.student) and not _teacher_can_access_student_course(
            request.user,
            self.student,
            self.course,
        ):
            return HttpResponseForbidden('Нет доступа к журналу студента')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        student = self.student
        course = self.course
        lessons = list(Lesson.objects.filter(course=course).order_by('date', 'id'))
        grades_map = {g.date: g for g in Grade.objects.filter(student=student, course=course)}
        attendance_map = {a.lesson_id: a for a in Attendance.objects.filter(student=student, lesson__course=course)}
        highlighted_grade_id = None
        highlighted_grade = None
        highlighted_lesson = None
        grade_id_raw = (self.request.GET.get('grade') or '').strip()
        if grade_id_raw.isdigit():
            highlighted_grade_id = int(grade_id_raw)
            highlighted_grade = Grade.objects.filter(
                pk=highlighted_grade_id,
                student=student,
                course=course,
            ).first()
            if highlighted_grade:
                highlighted_lesson = next((lesson for lesson in lessons if lesson.date == highlighted_grade.date), None)

        rows = []
        for lesson in lessons:
            grade = grades_map.get(lesson.date)
            attendance = attendance_map.get(lesson.id)
            rows.append(
                {
                    'lesson': lesson,
                    'grade': grade,
                    'grade_value': grade.value if grade else None,
                    'attendance': attendance.get_status_display() if attendance else '-',
                }
            )

        context['student'] = student
        context['course'] = course
        context['rows'] = rows
        context['highlighted_grade_id'] = highlighted_grade_id
        context['highlighted_grade'] = highlighted_grade
        context['highlighted_lesson'] = highlighted_lesson
        return context


def _is_admin(user):
    return user.is_superuser or user.groups.filter(name='Admin').exists()


def _is_parent(user):
    return user.groups.filter(name='Parent').exists()


def _chat_role(user):
    if Teacher.objects.filter(user=user).exists():
        return ChatMessage.SenderRole.TEACHER
    if Student.objects.filter(user=user).exists():
        return ChatMessage.SenderRole.STUDENT
    return ''


def _can_access_chat(user, chat):
    return chat.student.user_id == user.id or chat.teacher.user_id == user.id or _is_admin(user)


def _student_can_write_teacher(student, teacher):
    return Course.objects.filter(group=student.group, teacher=teacher).exists()


def _grade_context(grade):
    if not grade:
        return None
    lesson = Lesson.objects.filter(course=grade.course, date=grade.date).first()
    return {
        'grade': grade,
        'subject': grade.course.subject,
        'type': grade.get_grade_type_display(),
        'value': grade.value,
        'date': grade.date,
        'topic': lesson.topic if lesson else '',
        'url': (
            reverse('student_course_journal', kwargs={'student_id': grade.student_id, 'course_id': grade.course_id})
            + f'?grade={grade.id}#grade-{grade.id}'
        ),
    }


def _can_access_student_profile(user, student):
    if _is_admin(user):
        return True
    if student.user == user:
        return True
    if _is_parent(user) and Parent.objects.filter(user=user, children=student).exists():
        return True
    return False


def _teacher_can_manage_group(user, group):
    if _is_admin(user):
        return True
    return Teacher.objects.filter(user=user, courses__group=group).exists()


def _teacher_can_access_student_course(user, student, course):
    if _is_admin(user):
        return True
    return Teacher.objects.filter(user=user, courses=course, courses__group=student.group).exists()


def _build_username(first_name, last_name, group_id):
    base = slugify(f'{last_name}_{first_name}') or f'student_{group_id}'
    username = base
    seq = 1
    while User.objects.filter(username=username).exists():
        seq += 1
        username = f'{base}_{seq}'
    return username


@login_required
def import_students_to_group_view(request, group_id):
    from core.models import Group as StudyGroup

    study_group = get_object_or_404(StudyGroup, pk=group_id)
    if not _teacher_can_manage_group(request.user, study_group):
        from django.http import HttpResponseForbidden

        return HttpResponseForbidden('Нет доступа к импорту в эту группу')

    if request.method == 'POST':
        form = StudentImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded = form.cleaned_data['file']
            default_password = form.cleaned_data['default_password'] or 'student123'

            try:
                workbook = load_workbook(uploaded, read_only=True, data_only=True)
            except Exception:
                messages.error(request, 'Не удалось прочитать Excel файл. Проверьте, что это корректный .xlsx.')
                return redirect('group_students_import', group_id=study_group.id)

            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                messages.warning(request, 'Файл пустой.')
                return redirect('group_students_import', group_id=study_group.id)

            first_row = [str(c).strip().lower() if c is not None else '' for c in rows[0]]
            fio_col = 0
            has_header = False
            for idx, value in enumerate(first_row):
                if value in {'фио', 'fio', 'фио студента', 'student'}:
                    fio_col = idx
                    has_header = True
                    break

            data_rows = rows[1:] if has_header else rows

            student_role, _ = Group.objects.get_or_create(name='Student')
            created_users = 0
            created_students = 0
            moved_students = 0
            updated_names = 0
            skipped = 0

            for row in data_rows:
                if not row:
                    continue

                fio_value = ''
                if fio_col < len(row) and row[fio_col] is not None:
                    fio_value = str(row[fio_col]).strip()

                if not fio_value:
                    skipped += 1
                    continue

                fio_value = ' '.join(fio_value.replace(',', ' ').split())
                parts = fio_value.split()
                if len(parts) < 2:
                    skipped += 1
                    continue

                last_name = parts[0]
                first_name = ' '.join(parts[1:])
                username = _build_username(parts[1], last_name, study_group.id)

                user, user_created = User.objects.get_or_create(
                    username=username,
                    defaults={'first_name': first_name, 'last_name': last_name},
                )
                if user_created:
                    user.set_password(default_password)
                    user.save(update_fields=['password'])
                    created_users += 1
                else:
                    changed = False
                    if first_name and user.first_name != first_name:
                        user.first_name = first_name
                        changed = True
                    if last_name and user.last_name != last_name:
                        user.last_name = last_name
                        changed = True
                    if changed:
                        user.save(update_fields=['first_name', 'last_name'])
                        updated_names += 1

                user.groups.add(student_role)

                student, student_created = Student.objects.get_or_create(
                    user=user,
                    defaults={'group': study_group},
                )
                if student_created:
                    created_students += 1
                elif student.group_id != study_group.id:
                    student.group = study_group
                    student.save(update_fields=['group'])
                    moved_students += 1

            messages.success(
                request,
                f'Импорт завершен: users+{created_users}, students+{created_students}, перенесено {moved_students}, '
                f'обновлено ФИО {updated_names}, пропущено {skipped}.',
            )
            return redirect('course_list')
    else:
        form = StudentImportForm()

    return render(
        request,
        'users/import_students.html',
        {'form': form, 'study_group': study_group},
    )
